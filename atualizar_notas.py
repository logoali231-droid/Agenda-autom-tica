import json
import os
from datetime import datetime

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from calendar_sync import (
    get_calendar,
    criar_ou_atualizar_evento,
)


TOKEN = os.environ["CANVAS_TOKEN"]

BASE_URL = "https://pucpr.instructure.com/api/v1"

HEADERS = {
    "Authorization": f"Bearer {TOKEN}"
}

COURSES_FILE = "courses.json"
HISTORICO = "historico_atividades.json"
ARQUIVO_EXCEL = "Painel_Academico_PUCPR.xlsx"


def carregar_cursos():
    with open(COURSES_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def get_all_pages(url):
    dados = []

    while url:
        r = requests.get(url, headers=HEADERS)

        if r.status_code != 200:
            raise Exception(
                f"Erro {r.status_code}: {r.text}"
            )

        dados.extend(r.json())

        if "next" in r.links:
            url = r.links["next"]["url"]
        else:
            url = None

    return dados


def carregar_historico():
    try:
        with open(HISTORICO, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def salvar_historico(historico):
    with open(
        HISTORICO,
        "w",
        encoding="utf-8",
    ) as f:
        json.dump(
            historico,
            f,
            ensure_ascii=False,
            indent=2,
        )


def sincronizar_evento(
    calendar,
    curso_nome,
    course_id,
    assignment,
    score,
):
    due_at = assignment.get("due_at")

    if not due_at:
        return

    pontos = assignment.get("points_possible")
    html_url = assignment.get("html_url")

    try:
        inicio = datetime.fromisoformat(
            due_at.replace("Z", "+00:00")
        )
    except ValueError:
        print(
            "Data invalida:",
            curso_nome,
            assignment["name"],
        )
        return

    if score is not None:
        status_evento = "✅"
        status_texto = f"Concluida | Nota: {score}"
    else:
        status_evento = "📚"
        status_texto = "Pendente"

    descricao = (
        f"Disciplina: {curso_nome}\n\n"
        f"Vale: {pontos} pontos\n"
        f"Status: {status_texto}\n\n"
        f"Link Canvas:\n"
        f"{html_url or 'Sem link'}"
    )

    print(
        "ATUALIZANDO EVENTO:",
        f"{curso_nome} - {assignment['name']}",
    )

    criar_ou_atualizar_evento(
        calendar,
        f"{course_id}_{assignment['id']}",
        f"{status_evento} {curso_nome} - {assignment['name']}",
        descricao,
        inicio,
        html_url or "",
    )

    print("OK")


def main():
    courses = carregar_cursos()
    atividades_antigas = carregar_historico()

    ids_antigos = {
        item["id"]
        for item in atividades_antigas
    }

    novas_atividades = []

    resumo = []
    pendentes = []
    nao_contabilizadas = []

    calendar = get_calendar()

    writer = pd.ExcelWriter(
        ARQUIVO_EXCEL,
        engine="openpyxl",
    )

    for curso_nome, course_id in courses.items():

        print()
        print(f"Processando {curso_nome}")

        try:
            assignments = get_all_pages(
                f"{BASE_URL}/courses/"
                f"{course_id}/assignments"
                f"?per_page=100"
            )

            print(
                f"{curso_nome}: "
                f"{len(assignments)} atividades encontradas"
            )

            linhas = []

            total_obtido = 0.0
            total_maximo = 0.0
            total_futuro = 0.0

            for assignment in assignments:

                assignment_id = assignment["id"]
                nome = assignment["name"]

                # =========================================
                # HISTORICO DE NOVAS ATIVIDADES
                # =========================================

                if assignment_id not in ids_antigos:

                    novas_atividades.append(
                        {
                            "id": assignment_id,
                            "nome": nome,
                            "curso": curso_nome,
                        }
                    )

                    print(
                        "🆕 NOVA ATIVIDADE:",
                        curso_nome,
                        "-",
                        nome,
                    )

                # =========================================
                # IGNORAR ATIVIDADES FORA DA NOTA FINAL
                # =========================================

                if assignment.get(
                    "omit_from_final_grade",
                    False,
                ):
                    nao_contabilizadas.append(
                        {
                            "Disciplina": curso_nome,
                            "Atividade": nome,
                        }
                    )

                    continue

                pontos = assignment.get(
                    "points_possible"
                )

                due_at = assignment.get(
                    "due_at"
                )

                html_url = assignment.get(
                    "html_url"
                )

                if pontos is None or pontos <= 0:
                    continue

                # =========================================
                # BUSCAR NOTA
                # =========================================

                try:
                    response = requests.get(
                        f"{BASE_URL}/courses/"
                        f"{course_id}/assignments/"
                        f"{assignment_id}/submissions/self",
                        headers=HEADERS,
                    )

                    if response.status_code == 200:
                        submission = response.json()
                        score = submission.get("score")
                    else:
                        score = None

                except Exception:
                    score = None

                # =========================================
                # GOOGLE CALENDAR
                # =========================================

                sincronizar_evento(
                    calendar,
                    curso_nome,
                    course_id,
                    assignment,
                    score,
                )

                # =========================================
                # ATIVIDADE PENDENTE
                # =========================================

                if score is None:

                    total_futuro += float(pontos)

                    pendentes.append(
                        {
                            "Disciplina": curso_nome,
                            "Atividade": nome,
                            "Entrega": due_at,
                            "Valor Maximo": pontos,
                            "Link": html_url,
                        }
                    )

                    linhas.append(
                        {
                            "Atividade": nome,
                            "Entrega": due_at,
                            "Obtido": "",
                            "Maximo": pontos,
                            "%": "",
                            "Status": "Pendente",
                            "Link": html_url,
                        }
                    )

                    continue

                # =========================================
                # ATIVIDADE COM NOTA
                # =========================================

                percentual = round(
                    (
                        float(score)
                        / float(pontos)
                    ) * 100,
                    2,
                )

                total_obtido += float(score)
                total_maximo += float(pontos)

                if percentual >= 70:
                    status = "Bom"
                else:
                    status = "Atencao"

                linhas.append(
                    {
                        "Atividade": nome,
                        "Entrega": due_at,
                        "Obtido": score,
                        "Maximo": pontos,
                        "%": percentual,
                        "Status": status,
                        "Link": html_url,
                    }
                )

            # =============================================
            # CALCULOS
            # =============================================

            media = 0
            projecao = 0

            if total_maximo > 0:

                media = round(
                    (
                        total_obtido
                        / total_maximo
                    ) * 100,
                    2,
                )

                projecao = media

            if (
                total_futuro > 0
                and total_maximo > 0
            ):

                desempenho_atual = (
                    total_obtido
                    / total_maximo
                )

                nota_futura_esperada = (
                    total_futuro
                    * desempenho_atual
                )

                projecao = round(
                    (
                        total_obtido
                        + nota_futura_esperada
                    )
                    / (
                        total_maximo
                        + total_futuro
                    )
                    * 100,
                    2,
                )

            resumo.append(
                {
                    "Disciplina": curso_nome,
                    "Obtido": round(
                        total_obtido,
                        2,
                    ),
                    "Maximo Atual": round(
                        total_maximo,
                        2,
                    ),
                    "Pontos Pendentes": round(
                        total_futuro,
                        2,
                    ),
                    "Media Atual (%)": media,
                    "Projecao Final (%)": projecao,
                }
            )

            print(
                f"{curso_nome}: "
                f"{len(linhas)} linhas geradas"
            )

            df = pd.DataFrame(linhas)

            if not df.empty:
                df.to_excel(
                    writer,
                    sheet_name=curso_nome[:31],
                    index=False,
                )

        except Exception as e:

            print(
                f"ERRO EM {curso_nome}: {e}"
            )

            resumo.append(
                {
                    "Disciplina": curso_nome,
                    "Obtido": "ERRO",
                    "Maximo Atual": "",
                    "Pontos Pendentes": "",
                    "Media Atual (%)": "",
                    "Projecao Final (%)": str(e),
                }
            )

    # =============================================
    # ABAS EXTRAS
    # =============================================

    pd.DataFrame(
        nao_contabilizadas
    ).to_excel(
        writer,
        sheet_name="Nao_Contabilizadas",
        index=False,
    )

    pd.DataFrame(
        resumo
    ).to_excel(
        writer,
        sheet_name="Resumo",
        index=False,
    )

    pd.DataFrame(
        pendentes
    ).to_excel(
        writer,
        sheet_name="Pendentes",
        index=False,
    )

    writer.close()

    # =============================================
    # FORMATAR EXCEL
    # =============================================

    wb = load_workbook(
        ARQUIVO_EXCEL
    )

    for ws in wb.worksheets:

        for cell in ws[1]:

            cell.font = Font(
                bold=True
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="D9EAD3",
            )

        for col in ws.columns:

            tamanho = 0

            for cell in col:

                try:
                    tamanho = max(
                        tamanho,
                        len(
                            str(
                                cell.value
                            )
                        ),
                    )

                except Exception:
                    pass

            ws.column_dimensions[
                col[0].column_letter
            ].width = tamanho + 3

    wb.save(
        ARQUIVO_EXCEL
    )

    # =============================================
    # ATUALIZAR HISTORICO
    # =============================================

    historico = atividades_antigas.copy()

    historico.extend(
        novas_atividades
    )

    salvar_historico(
        historico
    )

    print()
    print("=" * 50)
    print("PLANILHA GERADA COM SUCESSO")
    print(ARQUIVO_EXCEL)
    print("=" * 50)


if __name__ == "__main__":
    main()
